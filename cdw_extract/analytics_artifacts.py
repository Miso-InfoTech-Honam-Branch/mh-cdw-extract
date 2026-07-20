"""분석 결과를 PNG·PDF·XLSX 파일로 생성하고 작업 상태와 콜백을 관리한다.

렌더러는 프로세스 메모리를 많이 사용하므로 동시 실행 수를 제한한다. 결과 파일은 부분 생성물이
완성본으로 보이지 않도록 임시 경로에 쓴 뒤 원자적으로 게시한다.
"""

from __future__ import annotations

import hashlib
import json
import math
import os
import re
import shutil
import threading
import time
import uuid
from collections import defaultdict
from contextlib import contextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterator

import requests

from cdw_extract.callback import post_json_callback
from cdw_extract.analytics import run_analytics_query
from cdw_extract.analytics_models import (
    AnalyticsArtifactRequest,
    AnalyticsQueryRequest,
    ArtifactFormat,
)
from cdw_extract.jobs import TERMINAL_STATES, create_job, load_job, normalize_job_id, update_job
from cdw_extract.user_dataset import file_sha256, safe_segment


ANALYSIS_ARTIFACT = "ANALYSIS_ARTIFACT"
READY = "READY"
_CHART_PALETTES: dict[str, tuple[str, ...]] = {
    "PROFESSIONAL": (
        "#2563EB", "#0F766E", "#7C3AED", "#EA580C", "#0891B2",
        "#D97706", "#DC2626", "#DB2777", "#65A30D", "#475569",
    ),
    "OCEAN": ("#075985", "#0284C7", "#06B6D4", "#14B8A6", "#2563EB", "#0F766E", "#4F46E5", "#64748B"),
    "WARM": ("#DC2626", "#EA580C", "#D97706", "#CA8A04", "#DB2777", "#9F1239", "#7C3AED", "#475569"),
    "ACCESSIBLE": ("#0072B2", "#E69F00", "#009E73", "#CC79A7", "#D55E00", "#56B4E9", "#F0E442", "#000000"),
}
_NUMBER_FORMATS = frozenset({"AUTO", "NUMBER", "COMPACT", "PERCENT", "CURRENCY_KRW"})
_active_guard = threading.Lock()
_active_cancellations: dict[str, threading.Event] = {}
_render_slots_guard = threading.Lock()
_render_slots: threading.BoundedSemaphore | None = None


class ArtifactCancelled(RuntimeError):
    pass


def _positive_number(name: str, default: float) -> float:
    raw = os.environ.get(name)
    if raw is None or not raw.strip():
        return default
    try:
        value = float(raw)
    except ValueError as exc:
        raise ValueError(f"{name} must be a number") from exc
    if value <= 0 or not math.isfinite(value):
        raise ValueError(f"{name} must be a positive finite number")
    return value


def _get_render_slots() -> threading.BoundedSemaphore:
    global _render_slots
    with _render_slots_guard:
        if _render_slots is None:
            configured = _positive_number("ANALYTICS_ARTIFACT_MAX_CONCURRENT", 1)
            if not configured.is_integer():
                raise ValueError("ANALYTICS_ARTIFACT_MAX_CONCURRENT must be a positive integer")
            capacity = int(configured)
            _render_slots = threading.BoundedSemaphore(capacity)
        return _render_slots


@contextmanager
def _artifact_render_slot(
    data_root: str | Path,
    request: AnalyticsArtifactRequest,
    cancellation: threading.Event,
) -> Iterator[None]:
    slots = _get_render_slots()
    timeout = _positive_number("ANALYTICS_ARTIFACT_QUEUE_TIMEOUT_SECONDS", 60)
    deadline = time.monotonic() + timeout
    acquired = False
    try:
        while not acquired:
            _check_cancelled(data_root, request, cancellation)
            remaining = deadline - time.monotonic()
            if remaining <= 0:
                raise TimeoutError(f"timed out waiting {timeout:g} seconds for an artifact render slot")
            acquired = slots.acquire(timeout=min(0.2, remaining))
        yield
    finally:
        if acquired:
            slots.release()


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _identity_segment(value: object, field_name: str) -> str:
    text = safe_segment(value, field_name)
    if any(ord(character) < 32 or character in '<>:"|?*' for character in text):
        raise ValueError(f"{field_name} contains characters that are not safe in a file path")
    if text.endswith((".", " ")):
        raise ValueError(f"{field_name} must not end with a dot or space")
    return text


def artifacts_root(data_root: str | Path) -> Path:
    return Path(data_root).expanduser().resolve() / "analysis-artifacts"


def artifact_root(data_root: str | Path, user_id: str, analysis_artifact_id: str) -> Path:
    return (
        artifacts_root(data_root)
        / _identity_segment(user_id, "userId")
        / _identity_segment(analysis_artifact_id, "analysisArtifactId")
    )


def artifact_manifest_path(data_root: str | Path, user_id: str, analysis_artifact_id: str) -> Path:
    return artifact_root(data_root, user_id, analysis_artifact_id) / "meta" / "manifest.json"


def artifact_relative_path(user_id: str, analysis_artifact_id: str, file_name: str) -> str:
    return (
        f"analysis-artifacts/{_identity_segment(user_id, 'userId')}/"
        f"{_identity_segment(analysis_artifact_id, 'analysisArtifactId')}/files/{file_name}"
    )


def _reservation_path(data_root: str | Path, user_id: str, analysis_artifact_id: str) -> Path:
    return (
        artifacts_root(data_root)
        / "_reservations"
        / _identity_segment(user_id, "userId")
        / f"{_identity_segment(analysis_artifact_id, 'analysisArtifactId')}.json"
    )


def _tombstone_path(data_root: str | Path, user_id: str, analysis_artifact_id: str) -> Path:
    return (
        artifacts_root(data_root)
        / "_tombstones"
        / _identity_segment(user_id, "userId")
        / f"{_identity_segment(analysis_artifact_id, 'analysisArtifactId')}.json"
    )


def _lock_path(data_root: str | Path, user_id: str, analysis_artifact_id: str) -> Path:
    return (
        artifacts_root(data_root)
        / "_locks"
        / _identity_segment(user_id, "userId")
        / f"{_identity_segment(analysis_artifact_id, 'analysisArtifactId')}.lock"
    )


@contextmanager
def _artifact_lock(data_root: str | Path, user_id: str, analysis_artifact_id: str) -> Iterator[None]:
    path = _lock_path(data_root, user_id, analysis_artifact_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a+b") as file:
        file.seek(0, os.SEEK_END)
        if file.tell() == 0:
            file.write(b"\0")
            file.flush()
        file.seek(0)
        if os.name == "nt":
            import msvcrt

            deadline = time.monotonic() + 60
            while True:
                try:
                    msvcrt.locking(file.fileno(), msvcrt.LK_NBLCK, 1)
                    break
                except OSError:
                    if time.monotonic() >= deadline:
                        raise TimeoutError("timed out locking analysis artifact")
                    time.sleep(0.01)
            try:
                yield
            finally:
                file.seek(0)
                msvcrt.locking(file.fileno(), msvcrt.LK_UNLCK, 1)
        else:
            import fcntl

            fcntl.flock(file.fileno(), fcntl.LOCK_EX)
            try:
                yield
            finally:
                fcntl.flock(file.fileno(), fcntl.LOCK_UN)


def _atomic_json(path: Path, value: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f"{path.name}.{os.getpid()}.{threading.get_ident()}.{uuid.uuid4().hex}.tmp")
    try:
        with temporary.open("w", encoding="utf-8") as file:
            json.dump(value, file, ensure_ascii=False, indent=2)
            file.flush()
            os.fsync(file.fileno())
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _safe_file_name(name: str, output_format: ArtifactFormat) -> str:
    stem = re.sub(r"[\\/:*?\"<>|\x00-\x1f]+", "_", str(name)).strip(" ._")[:180]
    while ".." in stem:
        stem = stem.replace("..", "_")
    if not stem:
        stem = "analysis"
    return f"{stem}.{output_format.value.lower()}"


def _accepted(request: AnalyticsArtifactRequest) -> dict:
    return {
        "jobId": request.job_id,
        "jobType": ANALYSIS_ARTIFACT,
        "requestId": request.request_id or request.analysis_artifact_id,
        "analysisArtifactId": request.analysis_artifact_id,
        "analysisId": request.analysis_id,
        "userId": request.user_id,
        "state": "ACCEPTED",
    }


def _spec_fingerprint(spec: dict[str, object]) -> str:
    # Boot creates per-dispatch chart request IDs. They correlate one execution
    # but do not change artifact content, so recovery may safely regenerate them.
    canonical = json.loads(json.dumps(spec, ensure_ascii=False))
    queries = canonical.get("queries") if isinstance(canonical, dict) else None
    if isinstance(queries, list):
        for item in queries:
            if isinstance(item, dict) and isinstance(item.get("query"), dict):
                item["query"].pop("requestId", None)
    encoded = json.dumps(canonical, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def prepare_analysis_artifact_job(
    request: AnalyticsArtifactRequest,
    data_root: str | Path,
) -> dict:
    normalized = request.model_copy(
        update={
            "job_id": normalize_job_id(request.job_id),
            "request_id": request.request_id or request.analysis_artifact_id,
            "user_id": _identity_segment(request.user_id, "userId"),
            "analysis_artifact_id": _identity_segment(request.analysis_artifact_id, "analysisArtifactId"),
        }
    )
    spec_hash = _spec_fingerprint(normalized.spec)
    with _artifact_lock(data_root, normalized.user_id, normalized.analysis_artifact_id):
        if _tombstone_path(data_root, normalized.user_id, normalized.analysis_artifact_id).exists():
            raise ValueError("analysis artifact was deleted and cannot be recreated with the same analysisArtifactId")
        final_manifest = artifact_manifest_path(data_root, normalized.user_id, normalized.analysis_artifact_id)
        if final_manifest.exists():
            existing_manifest = json.loads(final_manifest.read_text(encoding="utf-8"))
            if existing_manifest.get("jobId") != normalized.job_id:
                raise ValueError("analysisArtifactId already belongs to a different completed job")
        reservation_path = _reservation_path(data_root, normalized.user_id, normalized.analysis_artifact_id)
        reservation_created = False
        if reservation_path.exists():
            reservation = json.loads(reservation_path.read_text(encoding="utf-8"))
            expected = {
                "jobId": normalized.job_id,
                "userId": normalized.user_id,
                "analysisArtifactId": normalized.analysis_artifact_id,
                "analysisId": normalized.analysis_id,
                "format": normalized.format.value,
                "name": normalized.name,
                "specSha256": spec_hash,
            }
            if any(str(reservation.get(key)) != str(value) for key, value in expected.items()):
                raise ValueError("analysisArtifactId is already reserved by a different job")
        else:
            _atomic_json(
                reservation_path,
                {
                    "jobId": normalized.job_id,
                    "userId": normalized.user_id,
                    "analysisArtifactId": normalized.analysis_artifact_id,
                    "analysisId": normalized.analysis_id,
                    "format": normalized.format.value,
                    "name": normalized.name,
                    "specSha256": spec_hash,
                    "createdAt": utc_now(),
                },
            )
            reservation_created = True

        try:
            job, _created = create_job(
                data_root,
                {
                    "jobId": normalized.job_id,
                    "jobType": ANALYSIS_ARTIFACT,
                    "requestId": normalized.request_id,
                    "analysisArtifactId": normalized.analysis_artifact_id,
                    "analysisId": normalized.analysis_id,
                    "userId": normalized.user_id,
                    "format": normalized.format.value,
                    "name": normalized.name,
                    "specSha256": spec_hash,
                    "state": "ACCEPTED",
                },
            )
            expected_job = {
                "jobType": ANALYSIS_ARTIFACT,
                "analysisArtifactId": normalized.analysis_artifact_id,
                "analysisId": normalized.analysis_id,
                "userId": normalized.user_id,
                "format": normalized.format.value,
                "name": normalized.name,
                "specSha256": spec_hash,
            }
            if any(str(job.get(key)) != str(value) for key, value in expected_job.items()):
                raise ValueError("jobId is already assigned to a different analysis artifact")
        except Exception:
            if reservation_created:
                reservation_path.unlink(missing_ok=True)
            raise
    request.job_id = normalized.job_id
    request.request_id = normalized.request_id
    request.user_id = normalized.user_id
    request.analysis_artifact_id = normalized.analysis_artifact_id
    return _accepted(request)


def _check_cancelled(data_root: str | Path, request: AnalyticsArtifactRequest, event: threading.Event) -> None:
    if event.is_set() or _tombstone_path(data_root, request.user_id, request.analysis_artifact_id).exists():
        raise ArtifactCancelled("Analysis artifact generation was cancelled by deletion.")


def _normalized_display_options(value: object) -> dict[str, object]:
    raw = value if isinstance(value, dict) else {}
    palette_name = str(raw.get("palette") or "PROFESSIONAL").upper()
    if palette_name not in _CHART_PALETTES:
        palette_name = "PROFESSIONAL"

    number_format = str(raw.get("numberFormat") or "AUTO").upper()
    if str(raw.get("valueTransform") or "").upper() == "PERCENT_OF_TOTAL":
        number_format = "PERCENT"
    elif number_format not in _NUMBER_FORMATS:
        number_format = "AUTO"

    decimal_places = raw.get("decimalPlaces", 0)
    if isinstance(decimal_places, bool) or not isinstance(decimal_places, (int, float)):
        decimal_places = 0
    elif not math.isfinite(float(decimal_places)) or int(decimal_places) != decimal_places:
        decimal_places = 0
    decimal_places = max(0, min(3, int(decimal_places)))

    show_grid = raw.get("showGrid", True)
    if not isinstance(show_grid, bool):
        show_grid = True

    rotation_value = raw.get("axisLabelRotation", "AUTO")
    rotation = str(rotation_value).upper()
    if rotation not in {"AUTO", "0", "30", "45"}:
        rotation = "AUTO"
    return {
        "palette": palette_name,
        "colors": _CHART_PALETTES[palette_name],
        "numberFormat": number_format,
        "decimalPlaces": decimal_places,
        "showGrid": show_grid,
        "axisLabelRotation": rotation,
    }


def _chart_display_options(request: AnalyticsArtifactRequest, item: dict, chart_id: str) -> dict[str, object]:
    # Query options only contain execution concerns after Boot's allow-listing.
    # The original dashboard snapshot is therefore the source of truth for
    # presentation options, while the other sources keep legacy payloads safe.
    merged: dict[str, object] = {}
    query_payload = item.get("query")
    if isinstance(query_payload, dict) and isinstance(query_payload.get("options"), dict):
        merged.update(query_payload["options"])
    if isinstance(item.get("options"), dict):
        merged.update(item["options"])

    dashboard = request.spec.get("dashboard")
    if isinstance(dashboard, dict):
        if isinstance(dashboard.get("options"), dict):
            merged.update(dashboard["options"])
        charts = dashboard.get("charts")
        if isinstance(charts, list):
            matching_chart = next(
                (
                    chart for chart in charts
                    if isinstance(chart, dict) and str(chart.get("chartId") or "") == chart_id
                ),
                None,
            )
            if isinstance(matching_chart, dict) and isinstance(matching_chart.get("options"), dict):
                merged.update(matching_chart["options"])
    return _normalized_display_options(merged)


def _compiled_queries(request: AnalyticsArtifactRequest) -> list[dict]:
    queries = request.spec.get("queries")
    if not isinstance(queries, list) or not queries:
        raise ValueError("artifact spec.queries must contain compiled chart queries")
    if len(queries) > 50:
        raise ValueError("artifact spec.queries must not exceed 50 charts")
    compiled: list[dict] = []
    chart_ids: set[str] = set()
    for index, item in enumerate(queries):
        if not isinstance(item, dict) or not isinstance(item.get("query"), dict):
            raise ValueError(f"artifact spec.queries[{index}].query is required")
        chart_id = str(item.get("chartId") or f"chart-{index + 1}")
        if chart_id in chart_ids:
            raise ValueError(f"duplicate artifact chartId: {chart_id}")
        chart_ids.add(chart_id)
        query_payload = {**item["query"]}
        query_payload.setdefault("schemaVersion", 1)
        query_payload.setdefault("requestId", f"{request.request_id}:{chart_id}")
        query = AnalyticsQueryRequest.model_validate(query_payload)
        if query.source.user_id != request.user_id:
            raise ValueError(f"artifact query {chart_id} source.userId does not match artifact userId")
        layout = item.get("layout") if isinstance(item.get("layout"), dict) else {}
        compiled.append(
            {
                "chartId": chart_id,
                "title": str(item.get("title") or chart_id)[:255],
                "query": query,
                "displayOptions": _chart_display_options(request, item, chart_id),
                "layout": {
                    "x": max(0, min(48, int(layout.get("x") if layout.get("x") is not None else (index % 2) * 6))),
                    "y": max(0, min(96, int(layout.get("y") if layout.get("y") is not None else (index // 2) * 4))),
                    "w": max(1, min(24, int(layout.get("w") or 6))),
                    "h": max(1, min(24, int(layout.get("h") or 4))),
                },
            }
        )
    return compiled


def _font_configuration():
    from matplotlib.font_manager import FontProperties, findSystemFonts

    tokens = ("malgun", "notosanscjk", "noto sans cjk", "nanumgothic", "nanum gothic", "applegothic")
    for path in findSystemFonts(fontext="ttf"):
        normalized = Path(path).name.lower().replace("-", "").replace("_", "")
        if any(token.replace(" ", "") in normalized for token in tokens):
            return FontProperties(fname=path), Path(path).name, None
    return None, None, "No CJK font was found; Korean labels may use fallback glyphs."


def _labels(rows: list[dict], key: str) -> list[str]:
    return ["(null)" if row.get(key) is None else str(row.get(key)) for row in rows]


def _display_label(value: str) -> str:
    if re.match(r"^\d{4}-\d{2}-\d{2}T", value):
        return value[:7] if value[8:10] == "01" else value[:10]
    return value if len(value) <= 36 else value[:33] + "..."


def _numeric(value: object) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)) or not math.isfinite(float(value)):
        return 0.0
    return float(value)


def _format_number(value: object, options: dict[str, object], *, axis: bool = False) -> str:
    number = _numeric(value)
    if number == 0:
        number = 0.0
    decimal_places = int(options["decimalPlaces"])
    number_format = str(options["numberFormat"])
    compact = number_format == "COMPACT" or (number_format == "AUTO" and axis and abs(number) >= 1_000)
    suffix = ""
    scaled = number
    if compact:
        for divisor, unit in ((1_000_000_000_000, "조"), (100_000_000, "억"), (10_000, "만"), (1_000, "천")):
            if abs(number) >= divisor:
                scaled = number / divisor
                suffix = unit
                break
    formatted = f"{scaled:,.{decimal_places}f}"
    if number_format == "PERCENT":
        return f"{formatted}%"
    if number_format == "CURRENCY_KRW":
        unsigned = f"{abs(scaled):,.{decimal_places}f}"
        sign = "-" if scaled < 0 else ""
        return f"{sign}₩{unsigned}{suffix}"
    return f"{formatted}{suffix}"


def _axis_rotation(options: dict[str, object], category_count: int) -> int:
    configured = str(options["axisLabelRotation"])
    if configured != "AUTO":
        return int(configured)
    if category_count > 24:
        return 45
    if category_count > 12:
        return 30
    return 0


def _render_category(ax, chart_type: str, rows: list[dict], font, options: dict[str, object]) -> None:
    import numpy as np

    colors = list(options["colors"])
    series_values = list(dict.fromkeys(_labels(rows, "series"))) if any("series" in row for row in rows) else []
    categories = list(dict.fromkeys(_labels(rows, "category")))
    category_labels = [_display_label(value) for value in categories]
    if chart_type == "PIE":
        values = [_numeric(row.get("value")) for row in rows]
        decimals = int(options["decimalPlaces"])
        ax.pie(
            values,
            labels=[_display_label(value) for value in _labels(rows, "category")],
            autopct=f"%1.{decimals}f%%",
            colors=colors,
            textprops={"fontproperties": font},
        )
        return
    x = np.arange(len(categories))
    if series_values:
        width = 0.8 / max(1, len(series_values))
        lookup = {(str(row.get("category")), str(row.get("series"))): _numeric(row.get("value")) for row in rows}
        for index, series in enumerate(series_values):
            values = [lookup.get((category if category != "(null)" else "None", series), 0) for category in categories]
            positions = x + (index - (len(series_values) - 1) / 2) * width
            if chart_type == "LINE":
                ax.plot(x, values, marker="o", label=series)
            else:
                ax.bar(positions, values, width=width, label=series)
        ax.legend(prop=font, fontsize=8)
    else:
        values = [_numeric(row.get("value")) for row in rows]
        if chart_type == "LINE":
            ax.plot(x, values, marker="o")
        else:
            ax.bar(x, values)
    rotation = _axis_rotation(options, len(categories))
    ax.set_xticks(
        x,
        category_labels,
        rotation=rotation,
        ha="right" if rotation else "center",
        fontproperties=font,
        fontsize=7,
    )


def _render_scatter(ax, rows: list[dict], font, options: dict[str, object]) -> None:
    groups: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        groups[str(row.get("series") or "Data")].append(row)
    for label, group in groups.items():
        sizes = [max(12, min(600, _numeric(row.get("size")) * 5)) if "size" in row else 30 for row in group]
        ax.scatter([_numeric(row.get("x")) for row in group], [_numeric(row.get("y")) for row in group], s=sizes, alpha=0.65, label=label)
    if any("series" in row for row in rows):
        ax.legend(prop=font, fontsize=8)


def _render_boxplot(ax, rows: list[dict], font, options: dict[str, object]) -> None:
    from matplotlib.colors import to_rgba

    stats = [
        {
            "label": str(row.get("category") or "All"),
            "whislo": _numeric(row.get("min")),
            "q1": _numeric(row.get("q1")),
            "med": _numeric(row.get("median")),
            "q3": _numeric(row.get("q3")),
            "whishi": _numeric(row.get("max")),
            "fliers": [],
        }
        for row in rows
    ]
    colors = list(options["colors"])
    ax.bxp(
        stats,
        showfliers=False,
        patch_artist=True,
        boxprops={"facecolor": to_rgba(colors[0], 0.18), "edgecolor": colors[0], "linewidth": 1.4},
        medianprops={"color": colors[1], "linewidth": 1.6},
        whiskerprops={"color": colors[0]},
        capprops={"color": colors[0]},
    )
    rotation = _axis_rotation(options, len(stats))
    for label in ax.get_xticklabels():
        label.set_fontproperties(font)
        label.set_rotation(rotation)
        label.set_horizontalalignment("right" if rotation else "center")


def _render_funnel(ax, rows: list[dict], font, options: dict[str, object]) -> None:
    from matplotlib.patches import Polygon

    values = [max(0, _numeric(row.get("value"))) for row in rows]
    maximum = max(values, default=1) or 1
    count = max(1, len(rows))
    colors = list(options["colors"])
    for index, (row, value) in enumerate(zip(rows, values)):
        next_value = values[index + 1] if index + 1 < len(values) else value * 0.75
        top = 0.5 * value / maximum
        bottom = 0.5 * next_value / maximum
        y_top, y_bottom = count - index, count - index - 0.85
        ax.add_patch(
            Polygon(
                [(-top, y_top), (top, y_top), (bottom, y_bottom), (-bottom, y_bottom)],
                alpha=0.78,
                facecolor=colors[index % len(colors)],
            )
        )
        ax.text(
            0,
            (y_top + y_bottom) / 2,
            f"{row.get('category')}  {_format_number(value, options)}",
            ha="center",
            va="center",
            fontproperties=font,
            fontsize=8,
        )
    ax.set_xlim(-0.6, 0.6)
    ax.set_ylim(0, count + 0.2)
    ax.axis("off")


def _render_sankey(ax, rows: list[dict], font, options: dict[str, object]) -> None:
    from matplotlib.patches import Rectangle
    from matplotlib.path import Path as PlotPath
    from matplotlib.patches import PathPatch

    source_totals: dict[str, float] = defaultdict(float)
    target_totals: dict[str, float] = defaultdict(float)
    for row in rows:
        value = max(0, _numeric(row.get("value")))
        source_totals[str(row.get("source"))] += value
        target_totals[str(row.get("target"))] += value
    total = max(sum(source_totals.values()), 1)
    colors = list(options["colors"])

    def positions(values: dict[str, float]) -> dict[str, float]:
        cursor = 0.98
        result: dict[str, float] = {}
        for label, value in sorted(values.items(), key=lambda item: -item[1]):
            height = max(0.025, 0.86 * value / total)
            result[label] = cursor - height / 2
            cursor -= height + 0.02
        return result

    left, right = positions(source_totals), positions(target_totals)
    for label, y in left.items():
        ax.add_patch(Rectangle((0.03, y - 0.012), 0.025, 0.024, color=colors[0]))
        ax.text(0.065, y, label, va="center", ha="left", fontproperties=font, fontsize=7)
    for label, y in right.items():
        ax.add_patch(Rectangle((0.945, y - 0.012), 0.025, 0.024, color=colors[1 % len(colors)]))
        ax.text(0.935, y, label, va="center", ha="right", fontproperties=font, fontsize=7)
    for row in rows:
        source, target = str(row.get("source")), str(row.get("target"))
        if source not in left or target not in right:
            continue
        path = PlotPath([(0.055, left[source]), (0.35, left[source]), (0.65, right[target]), (0.945, right[target])], [PlotPath.MOVETO, PlotPath.CURVE4, PlotPath.CURVE4, PlotPath.CURVE4])
        width = max(0.8, 18 * max(0, _numeric(row.get("value"))) / total)
        ax.add_patch(PathPatch(path, fill=False, lw=width, alpha=0.28, color=colors[2 % len(colors)]))
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    ax.axis("off")


def _treemap_rectangles(values: list[tuple[str, float]], x: float, y: float, width: float, height: float, vertical: bool = True):
    total = sum(value for _, value in values) or 1
    cursor = x if vertical else y
    for label, value in values:
        ratio = value / total
        if vertical:
            rect = (cursor, y, width * ratio, height)
            cursor += width * ratio
        else:
            rect = (x, cursor, width, height * ratio)
            cursor += height * ratio
        yield label, value, rect


def _render_treemap(ax, rows: list[dict], font, options: dict[str, object]) -> None:
    from matplotlib.patches import Rectangle

    values = []
    for row in rows:
        levels = [str(value) for key, value in row.items() if key.startswith("level") and value is not None]
        values.append((" / ".join(levels) or "(null)", max(0, _numeric(row.get("value")))))
    colors = list(options["colors"])
    for index, (label, value, (x, y, width, height)) in enumerate(_treemap_rectangles(sorted(values, key=lambda item: -item[1]), 0, 0, 1, 1)):
        ax.add_patch(Rectangle((x, y), width, height, facecolor=colors[index % len(colors)], edgecolor="white"))
        if width > 0.06:
            ax.text(
                x + width / 2,
                y + height / 2,
                f"{label}\n{_format_number(value, options)}",
                ha="center",
                va="center",
                fontproperties=font,
                fontsize=7,
                clip_on=True,
            )
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    ax.axis("off")


def _render_chart(ax, chart_type: str, response, title: str, font, options: dict[str, object]) -> None:
    from matplotlib.ticker import FuncFormatter

    rows = response.rows
    ax.set_prop_cycle(color=list(options["colors"]))
    if not rows:
        ax.text(0.5, 0.5, "No data", ha="center", va="center", fontproperties=font)
    elif chart_type in {"BAR", "PIE", "LINE"}:
        _render_category(ax, chart_type, rows, font, options)
    elif chart_type == "SCATTER":
        _render_scatter(ax, rows, font, options)
    elif chart_type == "BOXPLOT":
        _render_boxplot(ax, rows, font, options)
    elif chart_type == "FUNNEL":
        _render_funnel(ax, rows, font, options)
    elif chart_type == "SANKEY":
        _render_sankey(ax, rows, font, options)
    elif chart_type == "TREEMAP":
        _render_treemap(ax, rows, font, options)
    else:
        raise ValueError(f"unsupported artifact chart type: {chart_type}")
    for line in response.reference_lines:
        ax.axhline(line.value, color=line.color or "#d62728", linestyle="--", linewidth=1, label=line.label)
    ax.set_title(title, fontproperties=font, fontsize=11)
    if chart_type not in {"PIE", "FUNNEL", "SANKEY", "TREEMAP"}:
        formatter = FuncFormatter(lambda value, _position: _format_number(value, options, axis=True))
        if chart_type == "SCATTER":
            ax.xaxis.set_major_formatter(formatter)
        ax.yaxis.set_major_formatter(formatter)
        if options["showGrid"]:
            ax.grid(True, alpha=0.2)
        else:
            ax.grid(False)
    else:
        ax.grid(False)
    for label in [*ax.get_xticklabels(), *ax.get_yticklabels()]:
        label.set_fontproperties(font)


def _build_figure(title: str, rendered: list[dict]):
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    font, font_name, font_warning = _font_configuration()
    max_right = max(item["layout"]["x"] + item["layout"]["w"] for item in rendered)
    max_bottom = max(item["layout"]["y"] + item["layout"]["h"] for item in rendered)
    figure = plt.figure(
        figsize=(min(32, max(10, max_right * 1.15)), min(32, max(6, max_bottom * 0.8))),
        dpi=140,
        facecolor="white",
    )
    warnings: list[str] = [font_warning] if font_warning else []
    for item in rendered:
        layout = item["layout"]
        left = layout["x"] / max_right + 0.018
        bottom = 1 - (layout["y"] + layout["h"]) / max_bottom + 0.035
        width = max(0.05, layout["w"] / max_right - 0.035)
        height = max(0.05, layout["h"] / max_bottom - 0.065)
        ax = figure.add_axes([left, bottom, width, height])
        try:
            _render_chart(
                ax,
                item["query"].chart_type.value,
                item["response"],
                item["title"],
                font,
                item["displayOptions"],
            )
        except Exception as exc:
            warning = f"Chart {item['chartId']} rendered as an error panel: {type(exc).__name__}: {exc}"
            warnings.append(warning)
            ax.clear()
            ax.text(0.5, 0.55, "Chart rendering failed", ha="center", va="center", color="#b91c1c", fontproperties=font)
            ax.text(0.5, 0.42, str(exc)[:240], ha="center", va="center", wrap=True, fontsize=8, fontproperties=font)
            ax.set_axis_off()
    figure.suptitle(title, x=0.01, y=0.997, ha="left", va="top", fontproperties=font, fontsize=15)
    return figure, font_name, warnings


def _write_xlsx(output_part: Path, preview_png: Path, title: str, rendered: list[dict]) -> None:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Font

    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    dashboard["A1"] = title
    dashboard["A1"].font = Font(size=16, bold=True)
    dashboard.add_image(Image(preview_png.as_posix()), "A3")
    used_names = {"Dashboard"}
    for index, item in enumerate(rendered, start=1):
        base = re.sub(r"[\\/*?:\[\]]", "_", item["title"])[:31] or f"Chart {index}"
        sheet_name = base
        suffix = 2
        while sheet_name in used_names:
            ending = f"_{suffix}"
            sheet_name = f"{base[:31-len(ending)]}{ending}"
            suffix += 1
        used_names.add(sheet_name)
        sheet = workbook.create_sheet(sheet_name)
        columns = item["response"].columns
        for column_index, column in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=column_index, value=column.label)
            cell.font = Font(bold=True)
        for row_index, row in enumerate(item["response"].rows, start=2):
            for column_index, column in enumerate(columns, start=1):
                sheet.cell(row=row_index, column=column_index, value=row.get(column.key))
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    workbook.save(output_part)


def _render_artifact(
    request: AnalyticsArtifactRequest,
    data_root: str | Path,
    event: threading.Event,
    staging_root: Path,
) -> tuple[Path, list[dict], dict]:
    compiled = _compiled_queries(request)
    rendered: list[dict] = []
    source_versions: dict[str, str] = {}
    for item in compiled:
        _check_cancelled(data_root, request, event)
        response = run_analytics_query(item["query"], data_root)
        rendered.append({**item, "response": response})
        source_versions[item["chartId"]] = response.source_version
    _check_cancelled(data_root, request, event)

    title = str(request.spec.get("title") or request.name)
    figure, font_name, render_warnings = _build_figure(title, rendered)
    file_name = _safe_file_name(request.name, request.format)
    output = staging_root / "files" / file_name
    output.parent.mkdir(parents=True, exist_ok=True)
    output_part = output.with_suffix(output.suffix + ".part")
    preview_png = staging_root / "dashboard-preview.png"
    try:
        if request.format == ArtifactFormat.PNG:
            figure.savefig(output_part, format="png", bbox_inches="tight", facecolor="white")
        elif request.format == ArtifactFormat.PDF:
            figure.savefig(output_part, format="pdf", bbox_inches="tight", facecolor="white")
        else:
            figure.savefig(preview_png, format="png", bbox_inches="tight", facecolor="white")
            _write_xlsx(output_part, preview_png, title, rendered)
        with output_part.open("r+b") as file:
            file.flush()
            os.fsync(file.fileno())
        os.replace(output_part, output)
    finally:
        output_part.unlink(missing_ok=True)
        preview_png.unlink(missing_ok=True)
        import matplotlib.pyplot as plt

        plt.close(figure)
    return output, rendered, {
        "fontName": font_name,
        "renderWarnings": render_warnings,
        "sourceVersions": source_versions,
    }


def _content_type(output_format: ArtifactFormat) -> str:
    return {
        ArtifactFormat.PNG: "image/png",
        ArtifactFormat.PDF: "application/pdf",
        ArtifactFormat.XLSX: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }[output_format]


def _callback_payload(request: AnalyticsArtifactRequest, status: str, manifest: dict | None = None, exc: Exception | None = None) -> dict:
    return {
        "jobId": request.job_id,
        "jobType": ANALYSIS_ARTIFACT,
        "requestId": request.request_id,
        "analysisArtifactId": request.analysis_artifact_id,
        "analysisId": request.analysis_id,
        "userId": request.user_id,
        "status": status,
        "fileName": manifest.get("fileName") if manifest else None,
        "relativePath": manifest.get("relativePath") if manifest else None,
        "contentType": manifest.get("contentType") if manifest else None,
        "sizeBytes": manifest.get("sizeBytes") if manifest else None,
        "sha256Checksum": manifest.get("sha256Checksum") if manifest else None,
        "sourceVersion": manifest.get("sourceVersion") if manifest else None,
        "errorCode": type(exc).__name__ if exc else None,
        "message": str(exc) if exc else None,
    }


def _post_callback(request: AnalyticsArtifactRequest, payload: dict) -> dict | None:
    callback = request.callback
    if callback is None:
        return None
    delivery = post_json_callback(
        {
            "url": callback.url,
            "headers": callback.headers,
            "timeoutSeconds": callback.timeout_seconds,
        },
        payload,
        operation="analysis artifact",
        post=requests.post,
    )
    if delivery is not None:
        delivery.pop("attempts", None)
        delivery["deliveredAt"] = utc_now()
    return delivery


def _deliver_callback(
    data_root: str | Path,
    request: AnalyticsArtifactRequest,
    payload: dict,
    cancellation: threading.Event | None = None,
) -> None:
    if request.callback is None:
        return
    last_error: Exception | None = None
    attempts = 0
    for attempt, delay in enumerate((0.0, 0.2, 0.5), start=1):
        if delay and cancellation is not None and cancellation.wait(delay):
            break
        if delay and cancellation is None:
            time.sleep(delay)
        attempts = attempt
        try:
            delivery = _post_callback(request, payload)
            if delivery:
                def mark_delivered(job: dict) -> dict:
                    current = {
                        **job,
                        "callbackDelivery": {**delivery, "attempt": attempt},
                        "callbackAttempts": attempt,
                    }
                    current.pop("callbackError", None)
                    return current

                update_job(
                    data_root,
                    request.job_id,
                    mark_delivered,
                )
            return
        except Exception as exc:
            last_error = exc
    if last_error is not None:
        try:
            update_job(
                data_root,
                request.job_id,
                lambda job: {
                    **job,
                    "callbackAttempts": attempts,
                    "callbackError": {
                        "errorCode": type(last_error).__name__,
                        "message": str(last_error),
                        "occurredAt": utc_now(),
                    },
                },
            )
        except Exception:
            pass


def run_analysis_artifact_job(request: AnalyticsArtifactRequest, data_root: str | Path) -> None:
    runner_token = uuid.uuid4().hex
    claimed = update_job(
        data_root,
        request.job_id,
        lambda current: {**current, "state": "RUNNING", "runnerToken": runner_token}
        if current.get("state") == "ACCEPTED"
        else current,
    )
    if claimed.get("state") == "SUCCESS":
        try:
            manifest = load_analysis_artifact_manifest(data_root, request.user_id, request.analysis_artifact_id)
            _deliver_callback(data_root, request, _callback_payload(request, READY, manifest))
        except Exception:
            pass
        return
    if claimed.get("runnerToken") != runner_token:
        return

    event = threading.Event()
    with _active_guard:
        _active_cancellations[request.job_id] = event
    staging_root = artifacts_root(data_root) / "_staging" / request.job_id
    shutil.rmtree(staging_root, ignore_errors=True)
    staging_root.mkdir(parents=True, exist_ok=False)
    try:
        with _artifact_render_slot(data_root, request, event):
            output, rendered, render_meta = _render_artifact(request, data_root, event, staging_root)
        _check_cancelled(data_root, request, event)
        file_name = output.name
        source_versions = render_meta.get("sourceVersions") or {}
        unique_source_versions = sorted(set(source_versions.values()))
        source_version = (
            unique_source_versions[0]
            if len(unique_source_versions) == 1
            else "sha256:"
            + hashlib.sha256(
                json.dumps(source_versions, sort_keys=True, separators=(",", ":")).encode("utf-8")
            ).hexdigest()
        )
        manifest = {
            "schemaVersion": 1,
            "jobId": request.job_id,
            "jobType": ANALYSIS_ARTIFACT,
            "requestId": request.request_id,
            "analysisArtifactId": request.analysis_artifact_id,
            "analysisId": request.analysis_id,
            "userId": request.user_id,
            "name": request.name,
            "format": request.format.value,
            "status": READY,
            "fileName": file_name,
            "relativePath": artifact_relative_path(request.user_id, request.analysis_artifact_id, file_name),
            "contentType": _content_type(request.format),
            "sizeBytes": output.stat().st_size,
            "sha256Checksum": file_sha256(output),
            "sourceVersion": source_version,
            "chartCount": len(rendered),
            **render_meta,
            "createdAt": utc_now(),
        }
        _atomic_json(staging_root / "meta" / "manifest.json", manifest)
        final_root = artifact_root(data_root, request.user_id, request.analysis_artifact_id)
        with _artifact_lock(data_root, request.user_id, request.analysis_artifact_id):
            _check_cancelled(data_root, request, event)
            if final_root.exists():
                existing = load_analysis_artifact_manifest(data_root, request.user_id, request.analysis_artifact_id)
                if existing.get("jobId") != request.job_id:
                    raise FileExistsError("analysis artifact target already exists")
                manifest = existing
                shutil.rmtree(staging_root, ignore_errors=True)
            else:
                final_root.parent.mkdir(parents=True, exist_ok=True)
                os.replace(staging_root, final_root)
        saved = update_job(
            data_root,
            request.job_id,
            lambda current: {
                **current,
                "state": "SUCCESS",
                "fileName": manifest["fileName"],
                "relativePath": manifest["relativePath"],
                "contentType": manifest["contentType"],
                "sizeBytes": manifest["sizeBytes"],
                "sha256Checksum": manifest["sha256Checksum"],
                "sourceVersion": manifest["sourceVersion"],
                "chartCount": manifest["chartCount"],
                "renderWarnings": manifest.get("renderWarnings") or [],
            },
        )
        if saved.get("state") != "SUCCESS":
            raise ArtifactCancelled("Artifact was cancelled before completion could be committed.")
        _deliver_callback(data_root, request, _callback_payload(request, READY, manifest), event)
    except ArtifactCancelled as exc:
        shutil.rmtree(staging_root, ignore_errors=True)
        update_job(
            data_root,
            request.job_id,
            lambda current: current if current.get("state") in TERMINAL_STATES else {**current, "state": "CANCELLED", "message": str(exc)},
        )
        _deliver_callback(data_root, request, _callback_payload(request, "CANCELLED", exc=exc))
    except Exception as exc:
        shutil.rmtree(staging_root, ignore_errors=True)
        if _tombstone_path(data_root, request.user_id, request.analysis_artifact_id).exists():
            cancelled = ArtifactCancelled("Analysis artifact generation was cancelled by deletion.")
            update_job(
                data_root,
                request.job_id,
                lambda current: current if current.get("state") in TERMINAL_STATES else {**current, "state": "CANCELLED", "message": str(cancelled)},
            )
            _deliver_callback(data_root, request, _callback_payload(request, "CANCELLED", exc=cancelled))
        else:
            update_job(
                data_root,
                request.job_id,
                lambda current: {**current, "state": "FAILED", "errorCode": type(exc).__name__, "message": str(exc)},
            )
            _deliver_callback(data_root, request, _callback_payload(request, "FAILED", exc=exc))
    finally:
        with _active_guard:
            _active_cancellations.pop(request.job_id, None)


def load_analysis_artifact_manifest(data_root: str | Path, user_id: str, analysis_artifact_id: str) -> dict:
    user_id = _identity_segment(user_id, "userId")
    analysis_artifact_id = _identity_segment(analysis_artifact_id, "analysisArtifactId")
    path = artifact_manifest_path(data_root, user_id, analysis_artifact_id)
    if not path.exists():
        raise FileNotFoundError(f"analysis artifact not found: {analysis_artifact_id}")
    manifest = json.loads(path.read_text(encoding="utf-8"))
    if manifest.get("userId") != user_id or manifest.get("analysisArtifactId") != analysis_artifact_id:
        raise ValueError("analysis artifact manifest identity mismatch")
    if manifest.get("status") != READY:
        raise ValueError("analysis artifact is not READY")
    return manifest


def analysis_artifact_download(data_root: str | Path, user_id: str, analysis_artifact_id: str) -> tuple[Path, dict]:
    manifest = load_analysis_artifact_manifest(data_root, user_id, analysis_artifact_id)
    root = Path(data_root).expanduser().resolve()
    relative = Path(str(manifest.get("relativePath") or ""))
    if relative.is_absolute():
        raise ValueError("analysis artifact path must be relative")
    output = (root / relative).resolve()
    expected = artifact_root(data_root, user_id, analysis_artifact_id).resolve() / "files" / str(manifest.get("fileName") or "")
    if output != expected.resolve() or not output.is_relative_to(root):
        raise ValueError("analysis artifact path does not match its canonical identity")
    if not output.is_file():
        raise FileNotFoundError("analysis artifact file is missing")
    if output.stat().st_size != int(manifest.get("sizeBytes") or -1):
        raise ValueError("analysis artifact size does not match its manifest")
    if file_sha256(output) != manifest.get("sha256Checksum"):
        raise ValueError("analysis artifact checksum does not match its manifest")
    return output, manifest


def delete_analysis_artifact(data_root: str | Path, user_id: str, analysis_artifact_id: str) -> dict:
    user_id = _identity_segment(user_id, "userId")
    analysis_artifact_id = _identity_segment(analysis_artifact_id, "analysisArtifactId")
    job_id: str | None = None
    with _artifact_lock(data_root, user_id, analysis_artifact_id):
        tombstone_path = _tombstone_path(data_root, user_id, analysis_artifact_id)
        existing_tombstone = (
            json.loads(tombstone_path.read_text(encoding="utf-8"))
            if tombstone_path.exists()
            else {}
        )
        job_id = existing_tombstone.get("jobId")
        reservation_path = _reservation_path(data_root, user_id, analysis_artifact_id)
        if reservation_path.exists():
            reservation = json.loads(reservation_path.read_text(encoding="utf-8"))
            job_id = reservation.get("jobId")
        manifest_path = artifact_manifest_path(data_root, user_id, analysis_artifact_id)
        if manifest_path.exists():
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            if manifest.get("userId") != user_id or manifest.get("analysisArtifactId") != analysis_artifact_id:
                raise ValueError("analysis artifact manifest identity mismatch")
            job_id = job_id or manifest.get("jobId")
        _atomic_json(
            tombstone_path,
            {
                "userId": user_id,
                "analysisArtifactId": analysis_artifact_id,
                "jobId": job_id,
                "deletedAt": existing_tombstone.get("deletedAt") or utc_now(),
            },
        )
        if job_id:
            with _active_guard:
                cancellation = _active_cancellations.get(str(job_id))
            if cancellation:
                cancellation.set()
            try:
                update_job(
                    data_root,
                    str(job_id),
                    lambda current: current if current.get("state") in TERMINAL_STATES else {**current, "state": "CANCELLED", "message": "Artifact was deleted."},
                )
            except FileNotFoundError:
                pass
        shutil.rmtree(artifact_root(data_root, user_id, analysis_artifact_id), ignore_errors=True)
        if job_id:
            shutil.rmtree(artifacts_root(data_root) / "_staging" / str(job_id), ignore_errors=True)
        reservation_path.unlink(missing_ok=True)
    return {"userId": user_id, "analysisArtifactId": analysis_artifact_id, "jobId": job_id, "state": "DELETED"}
