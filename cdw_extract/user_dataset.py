"""업로드 파일을 정규화해 안전한 사용자 데이터셋 Parquet으로 게시한다."""

from __future__ import annotations

import csv
import hashlib
import json
import os
import shutil
import uuid
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Callable

import requests

from .callback import callback_options as normalized_callback_options, post_json_callback
from .contracts import ResourceBudget
from .errors import ResourceLimitExceeded
from openpyxl import load_workbook

from .duck import connect, sql_literal

USER_DATASET_DIR = "user-datasets"
USER_DATASET_CONVERT = "USER_DATASET_CONVERT"


def utc_now() -> str:
    """UTC 현재 시각을 ISO 8601 문자열로 반환한다."""

    return datetime.now(timezone.utc).isoformat()


def safe_segment(value: object, field_name: str) -> str:
    """사용자 데이터셋 경로에 사용할 단일 식별자 조각을 검증한다."""

    text = str(value or "").strip()
    if not text:
        raise ValueError(f"{field_name} is required")
    if any(ch in text for ch in {"/", "\\", "\x00"}):
        raise ValueError(f"{field_name} must not contain path separators")
    if text in {".", ".."}:
        raise ValueError(f"{field_name} is invalid")
    return text


def user_dataset_root(data_root: str | Path) -> Path:
    """사용자 데이터셋 컬렉션의 루트 경로를 반환한다."""

    return Path(data_root) / USER_DATASET_DIR


def dataset_file_root(data_root: str | Path, user_id: str, user_dataset_id: str, user_dataset_file_id: str) -> Path:
    """사용자·데이터셋·파일 식별자에 해당하는 정규 저장 경로를 반환한다."""

    return (
        user_dataset_root(data_root)
        / safe_segment(user_id, "userId")
        / safe_segment(user_dataset_id, "userDatasetId")
        / "files"
        / safe_segment(user_dataset_file_id, "userDatasetFileId")
    )


def dataset_file_parquet_path(data_root: str | Path, user_id: str, user_dataset_id: str, user_dataset_file_id: str) -> Path:
    """게시된 사용자 데이터셋 Parquet 파일의 정규 경로를 반환한다."""

    return dataset_file_root(data_root, user_id, user_dataset_id, user_dataset_file_id) / "parquet" / "data.parquet"


def dataset_file_manifest_path(data_root: str | Path, user_id: str, user_dataset_id: str, user_dataset_file_id: str) -> Path:
    """사용자 데이터셋 파일 매니페스트의 정규 경로를 반환한다."""

    return dataset_file_root(data_root, user_id, user_dataset_id, user_dataset_file_id) / "meta" / "manifest.json"


def dataset_file_relative_path(user_id: str, user_dataset_id: str, user_dataset_file_id: str) -> str:
    """DATA_ROOT 기준 사용자 데이터셋 Parquet 파일 키를 만든다."""

    return (
        f"{USER_DATASET_DIR}/{safe_segment(user_id, 'userId')}/"
        f"{safe_segment(user_dataset_id, 'userDatasetId')}/files/"
        f"{safe_segment(user_dataset_file_id, 'userDatasetFileId')}/parquet/data.parquet"
    )


def dataset_file_manifest_relative_path(user_id: str, user_dataset_id: str, user_dataset_file_id: str) -> str:
    """DATA_ROOT 기준 사용자 데이터셋 매니페스트 키를 만든다."""

    return (
        f"{USER_DATASET_DIR}/{safe_segment(user_id, 'userId')}/"
        f"{safe_segment(user_dataset_id, 'userDatasetId')}/files/"
        f"{safe_segment(user_dataset_file_id, 'userDatasetFileId')}/meta/manifest.json"
    )


def load_dataset_file_manifest(data_root: str | Path, user_id: str, user_dataset_id: str, user_dataset_file_id: str) -> dict:
    """게시된 사용자 데이터셋 파일 매니페스트를 읽는다."""

    path = dataset_file_manifest_path(data_root, user_id, user_dataset_id, user_dataset_file_id)
    if not path.exists():
        raise FileNotFoundError(f"user dataset file manifest not found: {path}")
    return json.loads(path.read_text(encoding="utf-8"))


def save_dataset_file_manifest(
    data_root: str | Path,
    user_id: str,
    user_dataset_id: str,
    user_dataset_file_id: str,
    manifest: dict,
) -> dict:
    """사용자 데이터셋 매니페스트와 별도 스키마 문서를 저장한다."""

    path = dataset_file_manifest_path(data_root, user_id, user_dataset_id, user_dataset_file_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(path)
    schema_path = path.parent / "schema.json"
    schema_path.write_text(json.dumps(manifest.get("columns") or [], ensure_ascii=False, indent=2), encoding="utf-8")
    return manifest


def file_sha256(path: Path) -> str:
    """파일을 청크 단위로 읽어 SHA-256 체크섬을 계산한다."""

    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _publish_immutable_dataset_generation(
    staged_parquet_path: Path,
    final_file_root: Path,
    manifest: dict,
    matches_existing: Callable[[dict, Path, Path, dict], bool],
    conflict_message: str,
) -> tuple[dict, bool]:
    """Parquet과 메타데이터를 완성한 뒤 immutable generation 하나로 원자 게시한다."""

    if not staged_parquet_path.is_file():
        raise FileNotFoundError(f"staged Parquet file not found: {staged_parquet_path}")
    staged_file_root = staged_parquet_path.parent.parent
    final_parquet_path = final_file_root / "parquet" / "data.parquet"
    final_manifest_path = final_file_root / "meta" / "manifest.json"

    def matching_generation() -> dict | None:
        try:
            existing = json.loads(final_manifest_path.read_text(encoding="utf-8"))
            if not isinstance(existing, dict):
                return None
            return (
                existing
                if matches_existing(
                    existing,
                    final_parquet_path,
                    staged_parquet_path,
                    manifest,
                )
                else None
            )
        except (FileNotFoundError, OSError, json.JSONDecodeError):
            return None

    if final_file_root.exists():
        existing = matching_generation()
        if existing is not None:
            return existing, False
        raise FileExistsError(conflict_message)

    final_file_root.parent.mkdir(parents=True, exist_ok=True)
    publication_staging_root = final_file_root.parent / (
        f".{final_file_root.name}.{uuid.uuid4().hex}.staging"
    )
    try:
        publication_parquet_path = (
            publication_staging_root / "parquet" / "data.parquet"
        )
        publication_parquet_path.parent.mkdir(parents=True)
        try:
            # 같은 filesystem이면 hard link로 대용량 파일 복사를 피하고,
            # 다른 filesystem이거나 link를 지원하지 않으면 안전하게 복사한다.
            os.link(staged_parquet_path, publication_parquet_path)
        except OSError:
            shutil.copy2(staged_parquet_path, publication_parquet_path)

        # 독자가 Parquet만 있거나 manifest만 있는 중간 상태를 보지 않도록
        # 모든 파일을 최종 parent의 사설 sibling generation에서 완성한다.
        publication_meta_root = publication_staging_root / "meta"
        publication_meta_root.mkdir(parents=True)
        (publication_meta_root / "manifest.json").write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        (publication_meta_root / "schema.json").write_text(
            json.dumps(manifest.get("columns") or [], ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

        try:
            publication_staging_root.replace(final_file_root)
        except OSError:
            # 존재 여부 확인 뒤 다른 at-least-once 실행이 먼저 게시했으면
            # 호출 경로의 동일 generation 정책을 만족할 때만 멱등 성공으로 본다.
            existing = matching_generation()
            if existing is not None:
                return existing, False
            raise
    finally:
        shutil.rmtree(publication_staging_root, ignore_errors=True)

    # 기존 호출자가 staging root 소멸 여부로 신규 게시를 구분하므로
    # 최종 rename이 끝난 뒤 원본 private workspace도 정리한다.
    shutil.rmtree(staged_file_root, ignore_errors=True)
    if not final_manifest_path.is_file():
        raise FileNotFoundError(
            f"published user dataset manifest not found: {final_manifest_path}"
        )
    return manifest, True


def _extract_result_generation_matches(
    existing: dict,
    _final_parquet_path: Path,
    _staged_parquet_path: Path,
    requested: dict,
) -> bool:
    """기존 extract-result의 idempotency key와 checksum 정책을 적용한다."""

    return (
        existing.get("idempotencyKey") == requested.get("idempotencyKey")
        and existing.get("sha256Checksum") == requested.get("sha256Checksum")
        and existing.get("status") == "SUCCESS"
    )


def publish_dataset_file_artifact(
    staged_parquet_path: Path,
    data_root: str | Path,
    user_id: str,
    user_dataset_id: str,
    user_dataset_file_id: str,
    manifest: dict,
) -> dict:
    """Parquet, 매니페스트, 스키마를 하나의 USER_DATST 산출물로 원자 게시한다."""

    if not staged_parquet_path.is_file():
        raise FileNotFoundError(f"staged Parquet file not found: {staged_parquet_path}")
    final_file_root = dataset_file_root(
        data_root, user_id, user_dataset_id, user_dataset_file_id
    )
    normalized_manifest = {
        **manifest,
        "userId": safe_segment(user_id, "userId"),
        "userDatasetId": safe_segment(user_dataset_id, "userDatasetId"),
        "userDatasetFileId": safe_segment(user_dataset_file_id, "userDatasetFileId"),
        "path": dataset_file_relative_path(user_id, user_dataset_id, user_dataset_file_id),
        "manifestPath": dataset_file_manifest_relative_path(user_id, user_dataset_id, user_dataset_file_id),
        "fileType": "PARQUET",
        "sizeBytes": staged_parquet_path.stat().st_size,
        "sha256Checksum": file_sha256(staged_parquet_path),
        "status": "SUCCESS",
    }
    published_manifest, _ = _publish_immutable_dataset_generation(
        staged_parquet_path,
        final_file_root,
        normalized_manifest,
        _extract_result_generation_matches,
        f"user dataset result target already exists: {final_file_root}",
    )
    return published_manifest


def _conversion_generation_matches(
    existing: dict,
    final_parquet_path: Path,
    staged_parquet_path: Path,
    requested: dict,
) -> bool:
    """이미 게시된 변환 결과가 같은 job generation인지 확인한다."""

    if not final_parquet_path.is_file():
        return False
    # 동일 job 재시도는 완료 시각이 달라질 수 있지만 나머지 manifest와
    # 실제 Parquet 바이트는 모두 같아야 한다.
    existing_identity = {
        key: value for key, value in existing.items() if key != "createdAt"
    }
    requested_identity = {
        key: value for key, value in requested.items() if key != "createdAt"
    }
    return (
        existing_identity == requested_identity
        and final_parquet_path.stat().st_size == staged_parquet_path.stat().st_size
        and file_sha256(final_parquet_path) == file_sha256(staged_parquet_path)
    )


def _publish_dataset_conversion_generation(
    staged_parquet_path: Path,
    data_root: str | Path,
    user_id: str,
    user_dataset_id: str,
    user_dataset_file_id: str,
    manifest: dict,
) -> bool:
    """변환 결과를 immutable generation으로 게시하고 신규 게시 여부를 반환한다."""

    final_file_root = dataset_file_root(
        data_root, user_id, user_dataset_id, user_dataset_file_id
    )
    _, published = _publish_immutable_dataset_generation(
        staged_parquet_path,
        final_file_root,
        manifest,
        _conversion_generation_matches,
        f"user dataset conversion target already exists: {final_file_root}",
    )
    return published


def delete_user_dataset_file(data_root: str | Path, user_id: str, user_dataset_id: str, user_dataset_file_id: str) -> dict:
    """한 사용자 데이터셋 파일의 Parquet과 메타데이터를 함께 삭제한다."""

    root = dataset_file_root(data_root, user_id, user_dataset_id, user_dataset_file_id)
    shutil.rmtree(root, ignore_errors=True)
    return {
        "userId": safe_segment(user_id, "userId"),
        "userDatasetId": safe_segment(user_dataset_id, "userDatasetId"),
        "userDatasetFileId": safe_segment(user_dataset_file_id, "userDatasetFileId"),
        "state": "DELETED",
    }


def bool_option(value: object, default: bool = True) -> bool:
    """레거시 문자열을 포함한 옵션 값을 불리언으로 정규화한다."""

    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "t", "yes", "y", "on"}


def normalize_delimiter(value: str | None) -> str:
    """빈 구분자와 이스케이프된 탭을 실제 CSV 구분자로 변환한다."""

    if value is None or value == "":
        return ","
    if value == "\\t":
        return "\t"
    return value


def normalize_request_options(request: dict | None) -> dict:
    """중첩·평면 업로드 옵션을 호환 가능한 단일 요청 형태로 합친다."""

    request = request or {}
    options = request.get("options") if isinstance(request.get("options"), dict) else {}
    return {
        **request,
        "fileType": request.get("fileType") or options.get("fileType"),
        "headerYn": request.get("headerYn") if request.get("headerYn") is not None else options.get("header"),
        "delimiter": request.get("delimiter") or options.get("delimiter"),
        "fileEncoding": request.get("fileEncoding") or options.get("encoding"),
        "sheetName": request.get("sheetName") or options.get("sheetName"),
    }


def upload_suffix(filename: str | None, file_type: str | None = None) -> str:
    """명시 형식 또는 파일명에서 지원되는 업로드 확장자를 결정한다."""

    if file_type:
        normalized = file_type.strip().lower().lstrip(".")
        if normalized in {"csv", "xlsx", "parquet"}:
            return f".{normalized}"
        raise ValueError("fileType must be one of: CSV, XLSX, PARQUET")

    suffix = Path(filename or "").suffix.lower()
    if suffix not in {".csv", ".xlsx", ".parquet"}:
        raise ValueError("file type must be one of: csv, xlsx, parquet")
    return suffix


def copy_upload_file(upload: Any, output: Path) -> None:
    """업로드 스트림을 처음부터 작업 공간 파일로 복사한다."""

    output.parent.mkdir(parents=True, exist_ok=True)
    if hasattr(upload.file, "seek"):
        upload.file.seek(0)
    with output.open("wb") as target:
        shutil.copyfileobj(upload.file, target)


def unique_headers(values: list[Any]) -> list[str]:
    """비어 있거나 중복된 헤더를 안정적인 고유 열 이름으로 바꾼다."""

    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        text = str(value).strip() if value is not None else ""
        name = text or f"column_{index}"
        count = seen.get(name, 0)
        seen[name] = count + 1
        headers.append(name if count == 0 else f"{name}_{count + 1}")
    return headers


def csv_cell(value: Any) -> Any:
    """날짜·시간 셀을 CSV에서 손실 없이 읽을 수 있는 문자열로 바꾼다."""

    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return value


def padded_row(row: list[Any], width: int) -> list[Any]:
    """행을 지정한 열 수에 맞게 NULL로 채우거나 자른다."""

    values = list(row or [])
    values.extend([None] * (width - len(values)))
    return values[:width]


def generated_headers(width: int) -> list[str]:
    """헤더 없는 입력을 위한 순번 기반 열 이름을 만든다."""

    return [f"column_{index}" for index in range(1, width + 1)]


class _ByteBudgetWriter:
    def __init__(self, stream: Any, maximum_bytes: int | None, used_bytes: int = 0) -> None:
        self.stream = stream
        self.maximum_bytes = maximum_bytes
        self.used_bytes = used_bytes

    def write(self, value: str) -> int:
        encoded_size = len(value.encode("utf-8"))
        if self.maximum_bytes is not None and self.used_bytes + encoded_size > self.maximum_bytes:
            raise ResourceLimitExceeded(
                f"CSV/XLSX normalization exceeded resourceBudget.tempBytes={self.maximum_bytes}."
            )
        written = self.stream.write(value)
        self.used_bytes += encoded_size
        return written


def write_normalized_csv(
    rows: Any,
    output_path: Path,
    header: bool,
    max_temporary_bytes: int | None = None,
) -> None:
    """가변 폭 행을 디스크에 스풀해 고정 스키마 UTF-8 CSV로 정규화한다."""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    spool_path = output_path.with_name(f".{output_path.name}.{uuid.uuid4().hex}.rows")
    staged_output_path = output_path.with_name(
        f".{output_path.name}.{uuid.uuid4().hex}.normalized"
    )
    first_row: list[Any] | None = None
    max_width = 0
    try:
        # The widest row determines the final header width.  Spool rows to
        # disk so a large CSV/XLSX never has to be retained in Python memory.
        with spool_path.open("w", newline="", encoding="utf-8") as spool:
            spool_budget = _ByteBudgetWriter(spool, max_temporary_bytes)
            spool_writer = csv.writer(spool_budget)
            for row in rows:
                values = list(row or [])
                if first_row is None:
                    if not values or not any(value is not None and value != "" for value in values):
                        continue
                    first_row = values
                max_width = max(max_width, len(values))
                spool_writer.writerow([csv_cell(value) for value in values])
            spooled_bytes = spool_budget.used_bytes

        if first_row is None:
            raise ValueError("file has no rows")

        headers = (
            unique_headers(padded_row(first_row, max_width))
            if header
            else generated_headers(max_width)
        )
        with spool_path.open("r", newline="", encoding="utf-8") as spool, staged_output_path.open(
            "w", newline="", encoding="utf-8"
        ) as output:
            reader = csv.reader(spool)
            output_budget = _ByteBudgetWriter(output, max_temporary_bytes, spooled_bytes)
            writer = csv.writer(output_budget)
            writer.writerow(headers)
            for index, row in enumerate(reader):
                if header and index == 0:
                    continue
                writer.writerow(padded_row(row, max_width))
        # A failed budget check or encoding/write error must never leave a
        # partial normalized file at the caller-visible staging path.
        staged_output_path.replace(output_path)
    finally:
        spool_path.unlink(missing_ok=True)
        staged_output_path.unlink(missing_ok=True)


def csv_to_csv(
    input_path: Path,
    output_path: Path,
    delimiter: str,
    encoding: str | None,
    header: bool,
    max_temporary_bytes: int | None = None,
) -> None:
    """CSV 입력의 인코딩·구분자·헤더를 정규화된 CSV로 변환한다."""

    selected_encoding = encoding or "utf-8-sig"
    with input_path.open("r", newline="", encoding=selected_encoding) as file:
        reader = csv.reader(file, delimiter=delimiter)
        write_normalized_csv(reader, output_path, header, max_temporary_bytes)


def xlsx_to_csv(
    input_path: Path,
    output_path: Path,
    sheet_name: str | None = None,
    header: bool = True,
    max_temporary_bytes: int | None = None,
) -> None:
    """선택한 XLSX 시트를 읽기 전용으로 순회해 정규화된 CSV로 변환한다."""

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"sheetName not found: {sheet_name}")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        write_normalized_csv(rows, output_path, header, max_temporary_bytes)
    finally:
        workbook.close()


def source_sql_for_upload(input_path: Path, suffix: str) -> str:
    """정규화된 업로드 파일에 맞는 DuckDB 읽기 관계를 만든다."""

    path = sql_literal(input_path.as_posix())
    if suffix == ".parquet":
        return f"read_parquet({path})"
    return f"read_csv_auto({path}, HEADER=TRUE)"


def write_parquet_from_upload(
    input_path: Path,
    suffix: str,
    output_path: Path,
    data_root: str | Path | None = None,
    operation_id: object | None = None,
) -> None:
    """업로드 관계를 DuckDB로 읽어 Parquet 파일로 저장한다."""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    conn = connect(data_root, "user-dataset-convert", operation_id)
    try:
        conn.execute(
            f"COPY (SELECT * FROM {source_sql_for_upload(input_path, suffix)}) TO ? (FORMAT PARQUET)",
            [output_path.as_posix()],
        )
    finally:
        conn.close()


def parquet_columns(
    output_path: Path,
    data_root: str | Path | None = None,
    operation_id: object | None = None,
    connection=None,
) -> list[dict]:
    """Parquet의 실제 물리 스키마를 공개 열 메타데이터로 반환한다."""

    conn = connection or connect(data_root, "parquet-schema", operation_id)
    try:
        rows = conn.execute(
            f"DESCRIBE SELECT * FROM read_parquet({sql_literal(output_path.as_posix())})"
        ).fetchall()
    finally:
        if connection is None:
            conn.close()
    return [
        {
            "originalName": row[0],
            "name": row[0],
            "type": row[1],
            "ordinal": index,
            "nullable": True,
        }
        for index, row in enumerate(rows, start=1)
    ]


def parquet_row_count(
    output_path: Path,
    data_root: str | Path | None = None,
    operation_id: object | None = None,
    connection=None,
) -> int:
    """Parquet 파일의 전체 행 수를 DuckDB로 계산한다."""

    conn = connection or connect(data_root, "parquet-count", operation_id)
    try:
        return int(conn.execute("SELECT count(*) FROM read_parquet(?)", [output_path.as_posix()]).fetchone()[0])
    finally:
        if connection is None:
            conn.close()


def callback_options(request: dict) -> dict:
    """사용자 데이터셋 요청에서 콜백 전송 설정을 정규화한다."""

    return normalized_callback_options(request)


def post_callback(request: dict, payload: dict) -> dict | None:
    """사용자 데이터셋 변환 결과를 설정된 Boot 콜백으로 전송한다."""

    delivery = post_json_callback(
        callback_options(request),
        payload,
        operation="user dataset",
        post=requests.post,
    )
    if delivery is not None:
        delivery.pop("attempts", None)
    return delivery


def convert_user_dataset_file_from_path(
    input_upload_path: Path,
    data_root: str | Path,
    request: dict,
    *,
    budget: ResourceBudget | None = None,
    workspace: Path | None = None,
) -> dict:
    """작업 공간의 CSV·XLSX·Parquet을 검증된 사용자 데이터셋으로 변환한다."""

    request = normalize_request_options(request)
    user_id = safe_segment(request.get("userId"), "userId")
    user_dataset_id = safe_segment(request.get("userDatasetId"), "userDatasetId")
    user_dataset_file_id = safe_segment(request.get("userDatasetFileId"), "userDatasetFileId")
    original_file_name = request.get("originalFileName") or input_upload_path.name
    suffix = upload_suffix(original_file_name, request.get("fileType"))
    header = bool_option(request.get("headerYn"), default=True)
    delimiter = normalize_delimiter(request.get("delimiter"))
    encoding = request.get("fileEncoding")
    sheet_name = request.get("sheetName")
    job_id = str(request.get("jobId") or uuid.uuid4())
    request_id = str(request.get("requestId") or user_dataset_file_id)
    tmp_root = Path(workspace).resolve() if workspace is not None else user_dataset_root(data_root) / "_tmp" / job_id
    staged_parquet_path = tmp_root / "artifact" / "parquet" / "data.parquet"

    try:
        if suffix == ".xlsx":
            csv_path = tmp_root / "upload.csv"
            xlsx_to_csv(
                input_upload_path,
                csv_path,
                sheet_name=sheet_name,
                header=header,
                max_temporary_bytes=budget.temp_bytes if budget is not None else None,
            )
            input_path = csv_path
            input_suffix = ".csv"
        elif suffix == ".csv":
            csv_path = tmp_root / "upload.normalized.csv"
            csv_to_csv(
                input_upload_path,
                csv_path,
                delimiter=delimiter,
                encoding=encoding,
                header=header,
                max_temporary_bytes=budget.temp_bytes if budget is not None else None,
            )
            input_path = csv_path
            input_suffix = ".csv"
        else:
            input_path = input_upload_path
            input_suffix = suffix

        write_parquet_from_upload(input_path, input_suffix, staged_parquet_path, data_root, job_id)
        row_count = parquet_row_count(staged_parquet_path, data_root, job_id)
        if budget is not None and budget.row_limit is not None and row_count > budget.row_limit:
            raise ResourceLimitExceeded(
                f"User dataset output exceeded resourceBudget.rowLimit={budget.row_limit}."
            )
        output_size = staged_parquet_path.stat().st_size
        if budget is not None and budget.output_bytes is not None and output_size > budget.output_bytes:
            raise ResourceLimitExceeded(
                f"User dataset output exceeded resourceBudget.outputBytes={budget.output_bytes}."
            )
        if budget is not None:
            temporary_size = sum(
                path.stat().st_size for path in tmp_root.rglob("*") if path.is_file()
            )
            if temporary_size > budget.temp_bytes:
                raise ResourceLimitExceeded(
                    f"User dataset staging exceeded resourceBudget.tempBytes={budget.temp_bytes}."
                )
        columns = parquet_columns(staged_parquet_path, data_root, job_id)
        manifest = {
            "requestId": request_id,
            "jobId": job_id,
            "jobType": USER_DATASET_CONVERT,
            "userId": user_id,
            "userDatasetId": user_dataset_id,
            "userDatasetFileId": user_dataset_file_id,
            "originalFileName": original_file_name,
            "path": dataset_file_relative_path(user_id, user_dataset_id, user_dataset_file_id),
            "rowCount": row_count,
            "columns": columns,
            "fileType": suffix.lstrip(".").upper(),
            "headerYn": header,
            "delimiter": delimiter if suffix == ".csv" else None,
            "fileEncoding": encoding,
            "sheetName": sheet_name,
            "status": "SUCCESS",
            "createdAt": utc_now(),
        }
        # Parquet, manifest, schema를 사설 staging 세대 안에서 완성한 뒤
        # 디렉터리 하나를 원자적으로 게시한다. 같은 job의 재전달만 멱등 허용한다.
        _publish_dataset_conversion_generation(
            staged_parquet_path,
            data_root,
            user_id,
            user_dataset_id,
            user_dataset_file_id,
            manifest,
        )
        return {
            "requestId": request_id,
            "jobId": job_id,
            "jobType": USER_DATASET_CONVERT,
            "userId": user_id,
            "userDatasetId": user_dataset_id,
            "userDatasetFileId": user_dataset_file_id,
            "status": "SUCCESS",
            "rowCount": row_count,
            "columns": columns,
            "errorCode": None,
            "message": None,
        }
    finally:
        shutil.rmtree(tmp_root, ignore_errors=True)


def convert_user_dataset_file(upload: Any, data_root: str | Path, request: dict) -> dict:
    """HTTP 업로드를 작업 공간에 복사한 뒤 공통 경로 기반 변환을 실행한다."""

    request = normalize_request_options(request)
    job_id = str(request.get("jobId") or uuid.uuid4())
    tmp_root = user_dataset_root(data_root) / "_tmp" / job_id
    suffix = upload_suffix(getattr(upload, "filename", None), request.get("fileType"))
    upload_path = tmp_root / f"upload{suffix}"
    copy_upload_file(upload, upload_path)
    request = {**request, "jobId": job_id, "originalFileName": request.get("originalFileName") or getattr(upload, "filename", None)}
    return convert_user_dataset_file_from_path(upload_path, data_root, request)
