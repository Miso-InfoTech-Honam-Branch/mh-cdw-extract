from __future__ import annotations

import os
import tempfile
import threading
import uuid
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook
from pydantic import ValidationError

import app as api_app
import cdw_extract.analytics_artifacts as artifact_module
from cdw_extract.analytics import run_analytics_detail, run_analytics_query
from cdw_extract.analytics_artifacts import (
    analysis_artifact_download,
    artifact_root,
    delete_analysis_artifact,
    prepare_analysis_artifact_job,
    run_analysis_artifact_job,
)
from cdw_extract.analytics_models import (
    AnalyticsArtifactRequest,
    AnalyticsDetailRequest,
    AnalyticsQueryRequest,
)
from cdw_extract.jobs import load_job
from tests.test_analytics import (
    DATASET_ID,
    USER_ID,
    create_source,
    request_payload,
    run,
)


def calculated_field() -> dict:
    return {
        "id": "double_value",
        "name": "Double value",
        "dataType": "NUMBER",
        "formula": {
            "op": "MULTIPLY",
            "args": [
                {"op": "COALESCE", "args": [{"op": "COLUMN", "column": "value"}, {"op": "LITERAL", "value": 0}]},
                {"op": "LITERAL", "value": 2},
            ],
        },
    }


CHART_ENCODINGS = {
    "BAR": {"category": {"column": "category"}, "value": {"column": "value", "aggregation": "SUM"}},
    "PIE": {"category": {"column": "category"}, "value": {"aggregation": "COUNT"}},
    "LINE": {"category": {"column": "event_date", "timeGrain": "MONTH"}, "value": {"aggregation": "COUNT"}},
    "SCATTER": {"x": {"column": "x"}, "y": {"column": "y"}, "size": {"column": "size"}},
    "BOXPLOT": {"value": {"column": "value"}, "group": {"column": "box_group"}},
    "FUNNEL": {"stage": {"column": "stage"}, "value": {"aggregation": "COUNT"}},
    "SANKEY": {"source": {"column": "source_node"}, "target": {"column": "target_node"}, "value": {"aggregation": "COUNT"}},
    "TREEMAP": {"hierarchy": [{"column": "level1"}, {"column": "level2"}], "value": {"column": "value", "aggregation": "SUM"}},
}


def artifact_payload(
    output_format: str,
    analysis_artifact_id: str,
    *,
    all_charts: bool = False,
    callback: dict | None = None,
    display_options: dict | None = None,
) -> dict:
    chart_items = list(CHART_ENCODINGS.items()) if all_charts else [("BAR", CHART_ENCODINGS["BAR"])]
    queries = []
    dashboard_charts = []
    for index, (chart_type, encoding) in enumerate(chart_items):
        query = request_payload(chart_type, encoding)
        query["requestId"] = f"artifact-{analysis_artifact_id}-{chart_type.lower()}"
        chart_id = f"chart-{index + 1}"
        queries.append(
            {
                "chartId": chart_id,
                "title": chart_type,
                "query": query,
                "layout": {
                    "chartId": chart_id,
                    "x": (index % 2) * 6,
                    "y": (index // 2) * 4,
                    "w": 6,
                    "h": 4,
                },
            }
        )
        dashboard_charts.append({"chartId": chart_id, "options": dict(display_options or {})})
    payload = {
        "schemaVersion": 1,
        "jobId": str(uuid.uuid4()),
        "requestId": f"request-{analysis_artifact_id}",
        "analysisArtifactId": analysis_artifact_id,
        "analysisId": "analysis-v2",
        "userId": USER_ID,
        "name": f"Dashboard {analysis_artifact_id}",
        "format": output_format,
        "spec": {
            "specVersion": 2,
            "title": "Clinical dashboard",
            "dashboard": {"charts": dashboard_charts},
            "queries": queries,
        },
    }
    if callback:
        payload["callback"] = callback
    return payload


class AnalyticsDslV2Test(unittest.TestCase):
    def test_composable_filters_calculated_fields_fixed_bin_and_date_bucket(self):
        with tempfile.TemporaryDirectory() as data_root:
            create_source(data_root)
            payload = request_payload(
                "BAR",
                {
                    "category": {"column": "x", "bin": {"size": 2, "offset": 0}},
                    "value": {"derivedFieldId": "double_value", "aggregation": "SUM"},
                },
            )
            payload.update(
                {
                    "calculatedFields": [calculated_field()],
                    "globalFilters": [{"column": "active", "operator": "EQ", "value": True}],
                    "chartFilters": [{"column": "category", "operator": "NE", "value": "Z"}],
                    "interactionFilters": [
                        {
                            "column": "event_date",
                            "timeGrain": "MONTH",
                            "operator": "EQ",
                            "value": "2024-04-01",
                        }
                    ],
                    "referenceLines": [
                        {"id": "average", "type": "AVERAGE", "label": "Average"},
                        {"id": "target", "type": "TARGET", "value": 150},
                    ],
                }
            )
            response = run(data_root, payload)
            self.assertEqual([{"category": 4, "value": 200}], response.rows)
            self.assertEqual(3, response.metadata["appliedFilterCount"])
            self.assertEqual([200, 150], [line.value for line in response.reference_lines])

    def test_top_n_ratio_running_total_drilldown_and_comparisons(self):
        with tempfile.TemporaryDirectory() as data_root:
            create_source(data_root)
            top_payload = request_payload("BAR", CHART_ENCODINGS["BAR"])
            top_payload["topN"] = {
                "enabled": True,
                "count": 1,
                "by": "value",
                "direction": "DESC",
                "includeOthers": True,
            }
            top = run(data_root, top_payload)
            self.assertEqual(2, top.row_count)
            self.assertIn("Others", {row["category"] for row in top.rows})

            ratio_payload = request_payload("BAR", CHART_ENCODINGS["BAR"])
            ratio_payload["options"] = {"valueTransform": "PERCENT_OF_TOTAL"}
            ratio = run(data_root, ratio_payload)
            self.assertAlmostEqual(100, sum(float(row["value"]) for row in ratio.rows), places=7)

            running_payload = request_payload("LINE", CHART_ENCODINGS["LINE"])
            running_payload["options"] = {"valueTransform": "RUNNING_TOTAL"}
            running = run(data_root, running_payload)
            by_month = sorted(running.rows, key=lambda row: row["category"])
            self.assertEqual([2, 3, 4, 6, 7], [row["value"] for row in by_month])

            drill_payload = request_payload("BAR", CHART_ENCODINGS["BAR"])
            drill_payload["drilldown"] = {
                "fields": [{"column": "category"}, {"column": "series"}],
                "level": 1,
            }
            drill = run(data_root, drill_payload)
            self.assertEqual("series", drill.columns[0].label)
            self.assertFalse(drill.metadata["drilldown"]["canDrillDown"])

            series_payload = request_payload("BAR", CHART_ENCODINGS["BAR"])
            series_payload["comparison"] = {
                "enabled": True,
                "mode": "SERIES",
                "field": {"column": "series"},
                "offset": -1,
            }
            series = run(data_root, series_payload)
            self.assertTrue(all("series" in row for row in series.rows))

            period_payload = request_payload("LINE", CHART_ENCODINGS["LINE"])
            period_payload["comparison"] = {
                "enabled": True,
                "mode": "PREVIOUS_PERIOD",
                "periodUnit": "MONTH",
                "offset": -1,
            }
            period = run(data_root, period_payload)
            self.assertTrue({"previousValue", "change", "changeRate"}.issubset(period.rows[0]))
            january = min(period.rows, key=lambda row: row["category"])
            self.assertIsNone(january["previousValue"])
            self.assertTrue(any("previous available bucket" in warning for warning in period.warnings))

    def test_closed_expression_tree_rejects_raw_sql_and_bad_types(self):
        payload = request_payload("BAR", CHART_ENCODINGS["BAR"])
        payload["calculatedFields"] = [
            {"id": "unsafe", "name": "Unsafe", "formula": "value * 2; DROP TABLE x"}
        ]
        with self.assertRaises(ValidationError):
            AnalyticsQueryRequest.model_validate(payload)

        with tempfile.TemporaryDirectory() as data_root:
            create_source(data_root)
            invalid = request_payload(
                "BAR",
                {"category": {"column": "category"}, "value": {"derivedFieldId": "bad", "aggregation": "SUM"}},
            )
            invalid["calculatedFields"] = [
                {
                    "id": "bad",
                    "name": "Bad",
                    "dataType": "NUMBER",
                    "formula": {
                        "op": "ADD",
                        "args": [{"op": "COLUMN", "column": "category"}, {"op": "LITERAL", "value": 1}],
                    },
                }
            ]
            with self.assertRaisesRegex(ValueError, "numeric operands"):
                run(data_root, invalid)

    def test_boot_record_json_nulls_and_calculated_field_alias_are_compatible(self):
        with tempfile.TemporaryDirectory() as data_root:
            create_source(data_root)
            payload = request_payload(
                "BAR",
                {
                    "category": {"column": "category", "calculatedFieldId": None, "label": None, "aggregation": None, "timeGrain": None, "bin": None},
                    "value": {"column": None, "calculatedFieldId": "boot_calc", "label": None, "aggregation": "SUM", "timeGrain": None, "bin": None},
                    "hierarchy": [],
                },
            )
            payload.update(
                {
                    "calculatedFields": [
                        {
                            "id": "boot_calc",
                            "name": "Boot calc",
                            "dataType": "NUMBER",
                            "formula": {
                                "op": "MULTIPLY",
                                "column": None,
                                "value": None,
                                "args": [
                                    {
                                        "op": "COLUMN", "column": "value", "value": None,
                                        "args": [], "unit": None, "separator": None, "branches": [], "else": None,
                                    },
                                    {
                                        "op": "LITERAL", "column": None, "value": 2,
                                        "args": [], "unit": None, "separator": None, "branches": [], "else": None,
                                    },
                                ],
                                "unit": None,
                                "separator": None,
                                "branches": [],
                                "else": None,
                            },
                        }
                    ],
                    "globalFilters": [],
                    "chartFilters": [],
                    "interactionFilters": [],
                    "topN": None,
                    "drilldown": None,
                    "comparison": None,
                    "referenceLines": [],
                    "detailColumns": [],
                }
            )
            response = run(data_root, payload)
            self.assertGreater(response.row_count, 0)
            self.assertEqual("Boot calc", response.columns[1].label)

    def test_closed_case_concat_date_and_arithmetic_operators_execute(self):
        with tempfile.TemporaryDirectory() as data_root:
            create_source(data_root)
            source = request_payload("BAR", CHART_ENCODINGS["BAR"])["source"]
            calculated = [
                {
                    "id": "arithmetic",
                    "name": "Arithmetic",
                    "dataType": "NUMBER",
                    "formula": {
                        "op": "DIVIDE",
                        "args": [
                            {
                                "op": "SUBTRACT",
                                "args": [
                                    {
                                        "op": "ADD",
                                        "args": [
                                            {"op": "COLUMN", "column": "value"},
                                            {"op": "LITERAL", "value": 10},
                                        ],
                                    },
                                    {"op": "LITERAL", "value": 2},
                                ],
                            },
                            {"op": "LITERAL", "value": 2},
                        ],
                    },
                },
                {
                    "id": "joined",
                    "name": "Joined",
                    "dataType": "TEXT",
                    "formula": {
                        "op": "CONCAT",
                        "separator": "-",
                        "args": [
                            {"op": "COLUMN", "column": "category"},
                            {"op": "COLUMN", "column": "series"},
                        ],
                    },
                },
                {
                    "id": "active_label",
                    "name": "Active label",
                    "dataType": "TEXT",
                    "formula": {
                        "op": "CASE",
                        "branches": [
                            {
                                "when": {
                                    "op": "EQ",
                                    "args": [
                                        {"op": "COLUMN", "column": "active"},
                                        {"op": "LITERAL", "value": True},
                                    ],
                                },
                                "then": {"op": "LITERAL", "value": "Y"},
                            }
                        ],
                        "else": {"op": "LITERAL", "value": "N"},
                    },
                },
                {
                    "id": "event_year",
                    "name": "Event year",
                    "dataType": "NUMBER",
                    "formula": {
                        "op": "DATE_PART",
                        "unit": "YEAR",
                        "args": [{"op": "COLUMN", "column": "event_date"}],
                    },
                },
                {
                    "id": "zero_days",
                    "name": "Zero days",
                    "dataType": "NUMBER",
                    "formula": {
                        "op": "DATE_DIFF",
                        "unit": "DAY",
                        "args": [
                            {"op": "COLUMN", "column": "event_date"},
                            {"op": "COLUMN", "column": "event_date"},
                        ],
                    },
                },
            ]
            detail = AnalyticsDetailRequest.model_validate(
                {
                    "schemaVersion": 1,
                    "requestId": "operators",
                    "source": source,
                    "calculatedFields": calculated,
                    "detailColumns": [{"derivedFieldId": item["id"]} for item in calculated],
                    "limit": 1,
                }
            )
            row = run_analytics_detail(detail, data_root).rows[0]
            self.assertEqual(4.5, row["arithmetic"])
            self.assertEqual("A-S1", row["joined"])
            self.assertEqual("Y", row["active_label"])
            self.assertEqual(2024, row["event_year"])
            self.assertEqual(0, row["zero_days"])

    def test_detail_paging_uses_same_filters_and_calculated_fields(self):
        with tempfile.TemporaryDirectory() as data_root:
            create_source(data_root)
            source = request_payload("BAR", CHART_ENCODINGS["BAR"])["source"]
            request = AnalyticsDetailRequest.model_validate(
                {
                    "schemaVersion": 1,
                    "requestId": "detail-v2",
                    "source": source,
                    "calculatedFields": [calculated_field()],
                    "globalFilters": [{"column": "active", "operator": "EQ", "value": True}],
                    "detailColumns": [{"column": "category"}, {"derivedFieldId": "double_value"}],
                    "sorts": [{"field": "double_value", "direction": "DESC"}],
                    "offset": 0,
                    "limit": 2,
                }
            )
            response = run_analytics_detail(request, data_root)
            self.assertEqual(2, response.row_count)
            self.assertTrue(response.has_more)
            self.assertEqual([200, 14], [row["double_value"] for row in response.rows])


class AnalyticsArtifactTest(unittest.TestCase):
    def test_legacy_runner_reuses_callback_free_render_operation(self):
        with tempfile.TemporaryDirectory() as data_root:
            create_source(data_root)
            request = AnalyticsArtifactRequest.model_validate(
                artifact_payload("PNG", "artifact-shared-operation")
            )
            prepare_analysis_artifact_job(request, data_root)

            with patch(
                "cdw_extract.analytics_artifacts.render_analysis_artifact_operation",
                wraps=artifact_module.render_analysis_artifact_operation,
            ) as render_operation:
                run_analysis_artifact_job(request, data_root)

            render_operation.assert_called_once()
            self.assertTrue(render_operation.call_args.kwargs["check_tombstone"])
            path, manifest = analysis_artifact_download(
                data_root,
                USER_ID,
                "artifact-shared-operation",
            )
            self.assertTrue(path.read_bytes().startswith(b"\x89PNG\r\n\x1a\n"))
            self.assertEqual("READY", manifest["status"])

    def test_display_options_are_normalized_and_number_formats_match_frontend_contract(self):
        expected_first_colors = {
            "PROFESSIONAL": "#2563EB",
            "OCEAN": "#075985",
            "WARM": "#DC2626",
            "ACCESSIBLE": "#0072B2",
        }
        for palette, expected_color in expected_first_colors.items():
            with self.subTest(palette=palette):
                options = artifact_module._normalized_display_options({"palette": palette})
                self.assertEqual(palette, options["palette"])
                self.assertEqual(expected_color, options["colors"][0])

        invalid = artifact_module._normalized_display_options(
            {
                "palette": "unknown",
                "numberFormat": "SQL",
                "decimalPlaces": "many",
                "showGrid": "false",
                "axisLabelRotation": 90,
            }
        )
        self.assertEqual("PROFESSIONAL", invalid["palette"])
        self.assertEqual("AUTO", invalid["numberFormat"])
        self.assertEqual(0, invalid["decimalPlaces"])
        self.assertTrue(invalid["showGrid"])
        self.assertEqual("AUTO", invalid["axisLabelRotation"])

        cases = [
            ({"numberFormat": "NUMBER", "decimalPlaces": 2}, 1234.5, False, "1,234.50"),
            ({"numberFormat": "COMPACT", "decimalPlaces": 1}, 15000, False, "1.5만"),
            ({"numberFormat": "PERCENT", "decimalPlaces": 1}, 12.34, False, "12.3%"),
            ({"numberFormat": "CURRENCY_KRW", "decimalPlaces": 0}, -1234, False, "-₩1,234"),
            ({"numberFormat": "AUTO", "decimalPlaces": 1}, 1500, True, "1.5천"),
        ]
        for raw, value, axis, expected in cases:
            with self.subTest(number_format=raw["numberFormat"]):
                options = artifact_module._normalized_display_options(raw)
                self.assertEqual(expected, artifact_module._format_number(value, options, axis=axis))

    def test_dashboard_chart_options_drive_palette_axis_format_rotation_and_grid(self):
        from matplotlib.colors import to_rgba
        import matplotlib.pyplot as plt

        with tempfile.TemporaryDirectory() as data_root:
            create_source(data_root)
            display_options = {
                "palette": "OCEAN",
                "numberFormat": "CURRENCY_KRW",
                "decimalPlaces": 1,
                "showGrid": False,
                "axisLabelRotation": 45,
            }
            request = AnalyticsArtifactRequest.model_validate(
                artifact_payload("PNG", "artifact-display-options", display_options=display_options)
            )
            compiled = artifact_module._compiled_queries(request)
            self.assertEqual("OCEAN", compiled[0]["displayOptions"]["palette"])

            rendered = [{**compiled[0], "response": run_analytics_query(compiled[0]["query"], data_root)}]
            figure, _font_name, warnings = artifact_module._build_figure("Styled dashboard", rendered)
            try:
                self.assertFalse(any("rendered as an error panel" in warning for warning in warnings))
                axis = figure.axes[0]
                self.assertTrue(axis.patches)
                self.assertEqual(to_rgba("#075985"), axis.patches[0].get_facecolor())
                self.assertEqual("₩1,234.0", axis.yaxis.get_major_formatter()(1234, 0))
                self.assertTrue(all(label.get_rotation() == 45 for label in axis.get_xticklabels()))
                self.assertFalse(any(line.get_visible() for line in axis.yaxis.get_gridlines()))
            finally:
                plt.close(figure)

    def test_render_queue_timeout_finishes_as_failed_without_leaking_a_slot(self):
        with tempfile.TemporaryDirectory() as data_root:
            create_source(data_root)
            request = AnalyticsArtifactRequest.model_validate(
                artifact_payload("PNG", "artifact-render-timeout")
            )
            prepare_analysis_artifact_job(request, data_root)
            occupied_slot = threading.BoundedSemaphore(1)
            self.assertTrue(occupied_slot.acquire(blocking=False))

            try:
                with patch("cdw_extract.analytics_artifacts._render_slots", occupied_slot), patch.dict(
                    os.environ,
                    {"ANALYTICS_ARTIFACT_QUEUE_TIMEOUT_SECONDS": "0.05"},
                ):
                    run_analysis_artifact_job(request, data_root)
            finally:
                occupied_slot.release()

            job = load_job(data_root, request.job_id)
            self.assertEqual("FAILED", job["state"])
            self.assertEqual("TimeoutError", job["errorCode"])
            self.assertIn("artifact render slot", job["message"])

    def test_png_all_charts_pdf_and_xlsx_are_personal_ready_artifacts(self):
        with tempfile.TemporaryDirectory() as data_root:
            create_source(data_root)
            cases = [("PNG", True, b"\x89PNG\r\n\x1a\n"), ("PDF", True, b"%PDF"), ("XLSX", True, b"PK")]
            display_options = {
                "palette": "ACCESSIBLE",
                "numberFormat": "COMPACT",
                "decimalPlaces": 1,
                "showGrid": False,
                "axisLabelRotation": 30,
            }
            for output_format, all_charts, signature in cases:
                with self.subTest(output_format=output_format):
                    analysis_artifact_id = f"artifact-{output_format.lower()}"
                    request = AnalyticsArtifactRequest.model_validate(
                        artifact_payload(
                            output_format,
                            analysis_artifact_id,
                            all_charts=all_charts,
                            display_options=display_options,
                        )
                    )
                    accepted = prepare_analysis_artifact_job(request, data_root)
                    self.assertEqual("ACCEPTED", accepted["state"])
                    run_analysis_artifact_job(request, data_root)
                    path, manifest = analysis_artifact_download(data_root, USER_ID, analysis_artifact_id)
                    self.assertTrue(path.read_bytes().startswith(signature))
                    self.assertEqual(8, manifest["chartCount"])
                    self.assertTrue(manifest["sha256Checksum"])
                    self.assertGreater(manifest["sizeBytes"], 100)
                    if output_format == "XLSX":
                        workbook = load_workbook(path)
                        self.assertIn("Dashboard", workbook.sheetnames)
                        self.assertGreaterEqual(len(workbook["Dashboard"]._images), 1)

    def test_idempotency_download_integrity_and_deleted_artifact_cannot_resurrect(self):
        with tempfile.TemporaryDirectory() as data_root:
            create_source(data_root)
            payload = artifact_payload("PNG", "artifact-idempotent")
            request = AnalyticsArtifactRequest.model_validate(payload)
            prepare_analysis_artifact_job(request, data_root)
            prepare_analysis_artifact_job(request, data_root)
            regenerated = artifact_payload("PNG", "artifact-idempotent")
            regenerated["jobId"] = payload["jobId"]
            regenerated["requestId"] = payload["requestId"]
            regenerated["spec"]["queries"][0]["query"]["requestId"] = "regenerated-dispatch-id"
            prepare_analysis_artifact_job(AnalyticsArtifactRequest.model_validate(regenerated), data_root)
            run_analysis_artifact_job(request, data_root)
            run_analysis_artifact_job(request, data_root)
            path, _manifest = analysis_artifact_download(data_root, USER_ID, "artifact-idempotent")
            path.write_bytes(path.read_bytes() + b"tamper")
            with self.assertRaisesRegex(ValueError, "size does not match"):
                analysis_artifact_download(data_root, USER_ID, "artifact-idempotent")

            deleted = delete_analysis_artifact(data_root, USER_ID, "artifact-idempotent")
            self.assertEqual("DELETED", deleted["state"])
            self.assertFalse(artifact_root(data_root, USER_ID, "artifact-idempotent").exists())
            with self.assertRaisesRegex(ValueError, "deleted"):
                prepare_analysis_artifact_job(request, data_root)

    def test_delete_during_running_job_tombstones_before_late_publish(self):
        with tempfile.TemporaryDirectory() as data_root:
            create_source(data_root)
            request = AnalyticsArtifactRequest.model_validate(artifact_payload("PNG", "artifact-race"))
            prepare_analysis_artifact_job(request, data_root)
            started = threading.Event()
            release = threading.Event()

            def late_renderer(_request, _data_root, _event, staging_root: Path):
                started.set()
                self.assertTrue(release.wait(10))
                output = staging_root / "files" / "late.png"
                output.parent.mkdir(parents=True, exist_ok=True)
                output.write_bytes(b"\x89PNG\r\n\x1a\nlate")
                return output, [], {"fontName": None, "renderWarnings": [], "sourceVersions": {}}

            with patch("cdw_extract.analytics_artifacts._render_artifact", side_effect=late_renderer):
                thread = threading.Thread(target=run_analysis_artifact_job, args=(request, data_root))
                thread.start()
                self.assertTrue(started.wait(10))
                delete_analysis_artifact(data_root, USER_ID, "artifact-race")
                release.set()
                thread.join(10)
                self.assertFalse(thread.is_alive())
            self.assertFalse(artifact_root(data_root, USER_ID, "artifact-race").exists())
            self.assertEqual("CANCELLED", load_job(data_root, request.job_id)["state"])

    def test_callback_retries_three_times_and_job_supports_reconciliation(self):
        with tempfile.TemporaryDirectory() as data_root:
            create_source(data_root)
            request = AnalyticsArtifactRequest.model_validate(
                artifact_payload(
                    "PNG",
                    "artifact-callback",
                    callback={"url": "http://boot.invalid/callback", "timeoutSeconds": 1},
                )
            )
            prepare_analysis_artifact_job(request, data_root)

            class Response:
                text = "retry"

                def __init__(self, status_code):
                    self.status_code = status_code

            with patch(
                "cdw_extract.analytics_artifacts.requests.post",
                side_effect=[Response(503), Response(503), Response(204)],
            ) as post:
                run_analysis_artifact_job(request, data_root)
            self.assertEqual(3, post.call_count)
            callback_payload = post.call_args.kwargs["json"]
            self.assertEqual("ANALYSIS_ARTIFACT", callback_payload["jobType"])
            self.assertEqual("READY", callback_payload["status"])
            self.assertTrue(callback_payload["sourceVersion"].startswith("sha256:"))
            job = load_job(data_root, request.job_id)
            self.assertEqual("SUCCESS", job["state"])
            self.assertEqual(3, job["callbackAttempts"])
            self.assertIn("relativePath", job)
            self.assertIn("sha256Checksum", job)
            self.assertIn("sourceVersion", job)

    def test_routes_publish_detail_artifact_download_and_delete_contracts(self):
        routes = {getattr(route, "path", ""): route for route in api_app.app.routes}
        self.assertIn("/api/v1/analytics/detail", routes)
        self.assertEqual(200, routes["/api/v1/analytics/detail"].status_code)
        self.assertIn("/api/v1/analytics/artifacts", routes)
        self.assertEqual(202, routes["/api/v1/analytics/artifacts"].status_code)
        self.assertIn("/api/v1/analytics/artifacts/{userId}/{analysisArtifactId}/download", routes)
        self.assertIn("/api/v1/analytics/artifacts/{userId}/{analysisArtifactId}", routes)


if __name__ == "__main__":
    unittest.main()
